from openpyxl import load_workbook
from pathlib import Path

# xlsx 파일은 Data 폴더에 넣을 것.
# Path 객체에서 상대 경로는 working directory를 기준으로 해석됨.
# 어디에서 코드를 실행할지 확실치 않으므로 Data 폴더 찾을 때 현재 코드 위치를 기준으로 찾도록 해야 함.
filePath = Path(__file__).parents[1] / "Data" / "data.xlsx"
wb = load_workbook(filePath, data_only=True)
ws = wb.active 

# Count number of questions
questionNum = 0
for i in range(2, ws.max_row + 1):
    if isinstance(ws[f"A{i}"].value, int):
        questionNum += 1

# Check if each labels is valid
labelPass = 0
labelFail = 0

for i in range(2, ws.max_row + 1):
    if ws[f"C{i}"].value is not None: # 값이 있는 경우에만 실행
        answerType = ws[f"C{i}"].value
        answer = ws[f"D{i}"].value
        if answerType == "TF":
            if isinstance(answer, bool):
                labelPass += 1
            else:
                labelFail += 1
                print(f"{i}: TF label error")
        elif answerType == "MC":
            if isinstance(answer, int):
                labelPass += 1
            else:
                labelFail += 1
                print(f"{i}: MC label error")
        else:
            print(f"{i}: Invalid answer type")

wb.close()

print("<RESULT>")
print(f"Number of questions: {questionNum}")
print(f"Pass: {labelPass}, Fail: {labelFail}, Total: {labelPass + labelFail}")
